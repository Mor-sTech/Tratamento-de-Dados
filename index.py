import numpy as np 
import pandas as pd 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Border, Side, Font, PatternFill
from openpyxl.styles import Alignment

df = pd.read_csv('IMDB Top 250 Movies.csv', low_memory=False)
df.isnull().sum()

#Criando uma lista para armazenar os valores do box office
rev = []
for i in df['box_office'].values:
    rev.append(i)

#Removendo virgulas  
rev_c=[]
for i in rev:
    if ',' in str(i):
        j=i.replace(',','')
        rev_c.append(j)
    else:
        rev_c.append(i)  

##Lidando com a parte dos votos
rating = []
for i in df['rating']:
    rating.append(i)
rating  

v1 = []
for i in rating:
    if isinstance(i, str) and ',' in i:
        j=i.replace(',','')
        v1.append(j)
    else:
        v1.append(i)

df['rating'] = v1

df['rating']=df['rating'].astype(float)

df.dtypes

# Função para formatar valores em formato de número contábil do Excel
def format_accounting_number(value):
    try:
        value = float(value)  # Converter o valor para float
        return '{:,.2f}'.format(value)  # Formatar o valor com duas casas decimais e separadores de milhares
    except ValueError:
        return value  # Se não for possível converter para float, retorna o valor original

# Aplicar a formatação aos valores da coluna 'box_office'
df['box_office'] = df['box_office'].apply(format_accounting_number)

# Aplicar a formatação aos valores da coluna 'budget'
df['budget'] = df['budget'].apply(format_accounting_number)

# Salvar o DataFrame formatado em um arquivo Excel
df.to_excel('imdbTop250_tratado.xlsx', index=False)

# Renomeando várias colunas
df.rename(columns={'rank': 'Ranking', 'name': 'Titulo', 'year': 'Ano', 'rating': 'Avaliação', 'genre':'Genero','certificate' : 'Classificação-Indicativa', 'run_time':'Duração', 'tagline': 'Slogan', 'budget': 'Orçamento - (Dolar)', 'directors':'Diretores', 'casts' : 'Elenco', 'writers' : 'Escritores', 'box_office' : 'Faturamento - (Dolar)'}, inplace=True)
df.to_excel('imdbTop250_tratado.xlsx', index=False)

# Cria uma nova workbook
wb = Workbook()

# Selecionar a worksheet ativa
ws = wb.active

# Definir a largura das colunas
column_widths = {'A': 9, 'B': 50, 'C': 9, 'D': 9, 'E': 30, 'F': 30,'G': 10, 'H' : 50, 'I' : 25, 'J' : 25, 'K' : 50, 'L' : 45, 'M': 50} 

# Iterar sobre as colunas e ajustar a largura
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Converter DataFrame para o formato Excel e salvar no workbook
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Aplicar o formato de número contábil às células da coluna 'box_office'
for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=2, max_col=2):  # Coluna 'box_office' é a segunda coluna
    for cell in row:
        cell.style = 'Currency'  # Aplicar estilo de número contábil

# Centralizar e alinhar o texto no meio de todas as células
for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Criar um objeto de estilo de borda
border_style = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))

# Iterar sobre todas as células da planilha
for row in ws.iter_rows():
    for cell in row:
        cell.border = border_style

# Criar um objeto de estilo de fonte negrito
bold_font = Font(bold=True)

# Iterar sobre todas as células da linha 1 e aplicar o estilo de fonte negrito
for cell in ws[1]:
    cell.font = bold_font

# Criar um objeto de preenchimento com a cor desejada
fill = PatternFill(start_color="C9A0DC", end_color="C9A0DC", fill_type="solid")  # Cor rosa de exemplo

# Iterar sobre todas as células da linha 1 e aplicar o preenchimento de fundo
for cell in ws[1]:
    cell.fill = fill

# Salve as alterações no excel
wb.save('imdbTop250_tratado.xlsx')